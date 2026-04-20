#!/usr/bin/env python3
"""
导出脚本 - 将数据库内容导出为 Excel - 独立可运行版本

依赖：pip install openpyxl
用法：
  python3 export.py [--output report.xlsx] [--relevant-only]

数据库路径通过环境变量 DISCUSSION_DB_PATH 指定，默认 discussions.db
（需与 fetch_batch.py / annotate.py 使用同一数据库）
"""

import os
import sys
import sqlite3
import argparse
from datetime import datetime

DB_PATH = os.getenv("DISCUSSION_DB_PATH", "discussions.db")

COLUMNS = [
    "platform_id", "platform", "content_type", "title", "content",
    "url", "author", "subreddit", "score", "created_at",
    "search_keywords", "image_urls", "depth",
    "is_relevant", "summary", "advantages", "disadvantages", "tendency", "model",
]


def export_excel(output_path: str, relevant_only: bool = False):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"Error: database not found: {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    query = """
        SELECT d.platform_id, d.platform, d.content_type, d.title, d.content,
               d.url, d.author, d.subreddit, d.score, d.created_at,
               d.search_keywords, d.image_urls, d.depth,
               a.is_relevant, a.summary, a.advantages, a.disadvantages,
               a.tendency, a.model
        FROM discussions d
        LEFT JOIN annotations a ON d.platform_id = a.platform_id
    """
    if relevant_only:
        query += " WHERE a.is_relevant = 1"
    query += " ORDER BY d.created_at DESC"

    rows = conn.execute(query).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discussions"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col, key in enumerate(COLUMNS, 1):
            val = row[key]
            if key == "is_relevant" and val is not None:
                val = bool(val)
            ws.cell(row=row_idx, column=col, value=val)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f"Exported {len(rows)} rows -> {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export discussions DB to Excel")
    parser.add_argument("--output", default="annotated_report.xlsx", help="Output file path")
    parser.add_argument("--relevant-only", action="store_true", help="Only export relevant rows")
    args = parser.parse_args()
    export_excel(args.output, args.relevant_only)
