#!/usr/bin/env python3.11
"""
大模型数据市场调研 - Excel 生成脚本
用法: python3.11 generate_excel.py [--data JSON文件] [--output 输出目录]
"""
import json
import sys
import argparse
from datetime import date
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式常量 ─────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="F0F3F4")
THIN        = Side(border_style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_CONFIGS = {
    "招募动态": {
        "color": "E67E22",  # 橙色，突出核心招募数据
        "columns": ["序号","招募主体","招募任务类型","招募人员画像","招募发布时间","薪酬/单价信号","来源渠道","来源URL","内容摘要","采集日期"],
        "widths":  [5,     20,        28,             30,             14,             16,             14,         48,        40,         13    ],
    },
    "小红书&微信公众号": {
        "color": "C0392B",  # 红色
        "columns": ["序号","标题","作者/公众号","发布时间","命中关键词","内容摘要","分类标签","招募主体","招募任务类型","URL","采集日期"],
        "widths":  [5,     38,    18,            12,         20,         45,         16,         18,         22,            48,    13   ],
    },
    "网页搜索&行业媒体": {
        "color": "1A5276",
        "columns": ["序号","标题","来源平台","发布时间","涉及公司","关键数字","数据类型","内容摘要","URL","采集日期"],
        "widths":  [5,     40,    13,        12,         28,        20,        16,         52,        48,   13   ],
    },
    "厂商合作动态线索": {
        "color": "117A65",
        "columns": ["序号","大模型厂商","合作方/数据服务商","合作内容","数据类型","金额/规模","时间","训练阶段","来源","可信度","分析备注"],
        "widths":  [5,     14,          18,                  38,        18,         18,          11,    16,       18,     9,       35     ],
    },
}

# ── 分类标签规则 ─────────────────────────────────────────
LABEL_RULES = [
    (["招聘","岗位","内推","招募"],          "招聘/人才"),
    (["单价","薪资","工资","元/小时","元/h"], "薪酬/定价"),
    (["融资","A轮","B轮","C轮","投资"],      "融资/资本"),
    (["中标","采购","招标"],                  "政府/企业采购"),
    (["合成","DataFlow","自动生成"],          "合成数据"),
    (["视频","多模态","图文"],                "视频/多模态"),
    (["RLHF","SFT","post-training","后训练"], "Post-Training数据"),
    (["预训练","pre-training"],              "预训练数据"),
    (["数据飞轮","飞轮"],                    "数据飞轮"),
    (["架构","团队","组织","重组"],           "组织架构"),
    (["Scale AI","估值","对比"],             "行业分析"),
]

def auto_label(text: str) -> str:
    text_lower = (text or "").lower()
    for keywords, label in LABEL_RULES:
        if any(kw.lower() in text_lower for kw in keywords):
            return label
    return "行业生态"


def set_header(ws, columns, widths, fill_color):
    fill = PatternFill("solid", start_color=fill_color)
    for i, (col, w) in enumerate(zip(columns, widths), 1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = HEADER_FONT; c.fill = fill; c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def append_rows(ws, rows, start_row: int, row_height=45):
    for ri, row in enumerate(rows, start_row):
        bg = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = CELL_FONT; c.fill = bg; c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = row_height


def get_next_row(ws) -> int:
    """找到 Sheet 中第一个空行（跳过表头）"""
    for row in range(2, ws.max_row + 2):
        if ws.cell(row=row, column=1).value is None:
            return row
    return 2


def build_summary_sheet(ws4, recruit_count, xhs_count, web_count, coop_count):
    """构建汇总分析 Sheet"""
    hf4  = PatternFill("solid", start_color="4A235A")
    hf4b = PatternFill("solid", start_color="7D3C98")
    bold14 = Font(name="Arial", bold=True, size=14, color="4A235A")
    bold12w = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    bold10w = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    today = date.today().isoformat()
    ws4['A1'] = f"大模型数据市场信息汇总看板 — {today}"
    ws4['A1'].font = bold14
    ws4['A1'].alignment = Alignment(horizontal="center")
    ws4.merge_cells('A1:H1')
    ws4.row_dimensions[1].height = 36

    # 来源分布
    ws4['A3'] = "一、信息来源分布"
    ws4['A3'].font = bold12w; ws4['A3'].fill = hf4
    ws4.merge_cells('A3:D3')
    for ci, h in enumerate(["来源平台","信息条数","占比"], 1):
        c = ws4.cell(row=4, column=ci, value=h)
        c.font = bold10w; c.fill = hf4b; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    src_rows = [
        ("招募动态",           recruit_count, f"=B5/B9"),
        ("小红书&微信公众号",   xhs_count,     f"=B6/B9"),
        ("网页搜索/行业媒体",   web_count,     f"=B7/B9"),
        ("合作动态线索",        coop_count,    f"=B8/B9"),
        ("合计",               f"=SUM(B5:B8)", "100%"),
    ]
    for ri, row in enumerate(src_rows, 5):
        for ci, v in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", bold=(ri==9), size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
    for r in [5,6,7,8]:
        ws4.cell(row=r, column=3).number_format = '0.0%'

    for ci, w in enumerate([24, 12, 12, 12, 20, 20, 13, 28], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # 数据类型热度
    ws4['A10'] = "二、数据类型需求分布"
    ws4['A10'].font = bold12w; ws4['A10'].fill = hf4
    ws4.merge_cells('A10:E10')
    for ci, h in enumerate(["数据类型","提及次数","核心需求方","技术阶段","热度"], 1):
        c = ws4.cell(row=11, column=ci, value=h)
        c.font = bold10w; c.fill = hf4b; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    dtype_rows = [
        ("RLHF/后训练专家标注","—","百度/字节/阿里","Post-training","★★★★★"),
        ("预训练文本语料","—","数据堂/DeepSeek/智谱","Pre-training","★★★★"),
        ("视频/图文多模态","—","字节/阿里/中国移动","Pre-training","★★★★★"),
        ("语音识别/合成","—","海天瑞声/科大讯飞","Pre-training","★★★"),
        ("垂直领域专家标注","—","阿里(政治)/百度(金融)","Post-training","★★★★"),
        ("合成/自动生成数据","—","DataFlow/字节/DeepSeek","RL/Mid-training","★★★★"),
        ("代码/数学/推理数据","—","DeepSeek/字节Seed/智谱","Post-training/RL","★★★★"),
        ("政府/企业采购数据集","—","中国移动/中国电信","全阶段","★★★"),
    ]
    for ri, row in enumerate(dtype_rows, 12):
        bg = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for ci, v in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = CELL_FONT; c.fill = bg; c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # 厂商战略对比
    ws4['A21'] = "三、头部厂商数据战略对比矩阵"
    ws4['A21'].font = bold12w; ws4['A21'].fill = hf4
    ws4.merge_cells('A21:H21')
    for ci, h in enumerate(["厂商","数据战略","核心数据类型","获取方式","主攻阶段","规模信号","竞争力","战略特点"], 1):
        c = ws4.cell(row=22, column=ci, value=h)
        c.font = bold10w; c.fill = hf4b; c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    strat_rows = [
        ("百度","自建为主+平台生态","RLHF/金融/政务垂类","海口自建标注基地+千帆平台","Post-training","2025中标第二，8898万","★★★★★","全栈布局，芯片-云-模型-数据闭环"),
        ("字节跳动","高端化+视频多模态","L4专家标注/视频数据","与华胜天成合作+flow团队","Pre+Post均强","视频大规模招聘data岗","★★★★★","视频数据天然优势；L4高端标注质量导向"),
        ("阿里巴巴","工具化+生态投资","多模态RLHF/垂直专家","PAI工具平台+投资博登智能","Post-training重点","通义专家招募持续","★★★★","平台生态最大；数据工具化领先"),
        ("科大讯飞","垂类应用+标注运营","教育/医疗/金融语音","自建标注基地+大量外采","应用数据驱动","2025标王，2.32亿","★★★★","语音垂类壁垒最深；中标市场第一"),
        ("DeepSeek","RL范式+高效少量","代码/数学/推理强化","社区众包+精准少量标注","RL为核心","开源后催生大量外部需求","★★★★","以小博大，数据效率最高"),
        ("智谱AI","供给/自用双角色","文本推理/图文理解","参与政府采购+内部","全阶段","中标中移动688万","★★★","既是甲方也是乙方；学术背景强"),
        ("腾讯（混元）","架构重组后待观察","全品类","内部为主，外部合作少","全阶段","2025年12月重组","★★★","微信数据潜在优势待释放"),
    ]
    for ri, row in enumerate(strat_rows, 23):
        bg = ALT_FILL if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for ci, v in enumerate(row, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = CELL_FONT; c.fill = bg; c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws4.row_dimensions[ri].height = 50


def create_or_update_excel(data: dict, output_dir: str = "~/Downloads") -> str:
    """
    创建或更新 Excel 文件。

    data 结构：
    {
        "招募动态": [...],
        "小红书&微信公众号": [...],
        "网页搜索&行业媒体": [...],
        "厂商合作动态线索": [...],
    }
    每个列表的元素是与 Sheet 列对应的 tuple/list。
    """
    today = date.today().isoformat()
    output_path = Path(output_dir).expanduser() / f"大模型数据市场调研_{today}.xlsx"

    if output_path.exists():
        wb = load_workbook(str(output_path))
        print(f"追加到已有文件: {output_path}")
    else:
        wb = Workbook()
        wb.active.title = "招募动态"

        # 初始化各 Sheet 表头
        for sheet_name, cfg in SHEET_CONFIGS.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name) if sheet_name != "招募动态" else wb["招募动态"]
            else:
                ws = wb[sheet_name]
            set_header(ws, cfg["columns"], cfg["widths"], cfg["color"])

        # 汇总分析 Sheet
        if "汇总分析" not in wb.sheetnames:
            ws4 = wb.create_sheet("汇总分析")
        else:
            ws4 = wb["汇总分析"]

    # 追加数据
    for sheet_name, rows in data.items():
        if sheet_name not in wb.sheetnames:
            cfg = SHEET_CONFIGS[sheet_name]
            ws = wb.create_sheet(sheet_name)
            set_header(ws, cfg["columns"], cfg["widths"], cfg["color"])
        else:
            ws = wb[sheet_name]

        # 自动填充序号和分类标签
        next_row = get_next_row(ws)
        seq_start = next_row - 1  # 已有数据行数

        processed_rows = []
        for i, row in enumerate(rows, seq_start + 1):
            row_list = list(row)
            row_list[0] = i  # 序号自动递增

            # 小红书&微信公众号 Sheet 自动打分类标签（第7列）
            if sheet_name == "小红书&微信公众号" and len(row_list) >= 7 and not row_list[6]:
                text = " ".join(str(x) for x in row_list[1:6])
                row_list[6] = auto_label(text)

            processed_rows.append(tuple(row_list))

        append_rows(ws, processed_rows, next_row)

    # 更新汇总分析
    ws4 = wb["汇总分析"] if "汇总分析" in wb.sheetnames else wb.create_sheet("汇总分析")
    ws4.delete_rows(1, ws4.max_row)  # 清空后重建（保持最新统计）
    recruit_count = (wb["招募动态"].max_row - 1) if "招募动态" in wb.sheetnames else 0
    xhs_count  = (wb["小红书&微信公众号"].max_row - 1) if "小红书&微信公众号" in wb.sheetnames else 0
    web_count  = (wb["网页搜索&行业媒体"].max_row - 1) if "网页搜索&行业媒体" in wb.sheetnames else 0
    coop_count = (wb["厂商合作动态线索"].max_row - 1) if "厂商合作动态线索" in wb.sheetnames else 0
    build_summary_sheet(ws4, recruit_count, xhs_count, web_count, coop_count)

    wb.save(str(output_path))
    print(f"Excel 已保存: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成大模型数据调研 Excel")
    parser.add_argument("--data",   default=None, help="JSON 数据文件路径")
    parser.add_argument("--output", default="~/Downloads", help="输出目录")
    args = parser.parse_args()

    if args.data:
        with open(args.data, encoding="utf-8") as f:
            data = json.load(f)
    else:
        # 示例空数据
        data = {"小红书": [], "网页搜索&行业媒体": [], "厂商合作动态线索": []}

    path = create_or_update_excel(data, args.output)
    print(f"完成: {path}")
